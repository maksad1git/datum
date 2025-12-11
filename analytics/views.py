from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView
from django.db.models import Avg, Count, Sum, Min, Max
from django.utils import timezone
from django.http import HttpResponse
from django.template.loader import render_to_string
from datetime import timedelta
import json

from .models import Dashboard, Report, ReportTemplate, FilterPreset
from .forms import DashboardForm, ReportForm, ReportTemplateForm, FilterPresetForm
from visits.models import Visit, Observation
from coefficients.models import Coefficient
from geo.models import Country, Region, City, District, Channel, Outlet
from catalog.models import AttributeGroup

# WeasyPrint будет импортирован только при необходимости
WEASYPRINT_AVAILABLE = False


# ============================================================================
# Dashboard Views
# ============================================================================

class DashboardListView(LoginRequiredMixin, ListView):
    model = Dashboard
    template_name = 'analytics/dashboard_list.html'
    context_object_name = 'dashboards'
    paginate_by = 25

    def get_queryset(self):
        return Dashboard.objects.select_related('owner')


class DashboardDetailView(LoginRequiredMixin, DetailView):
    model = Dashboard
    template_name = 'analytics/dashboard_detail.html'
    context_object_name = 'dashboard'

    def get_queryset(self):
        return Dashboard.objects.select_related('owner').prefetch_related('shared_with')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Получить фильтры из GET параметров
        filters = self.get_filters()

        # Парсить widgets_config JSON и вычислять метрики
        config = {}
        if self.object.widgets_config:
            try:
                config = json.loads(self.object.widgets_config) if isinstance(self.object.widgets_config, str) else self.object.widgets_config
            except:
                config = {}

        widgets = []

        for widget_config in config.get('widgets', []):
            try:
                widget_type = widget_config.get('type')

                # Метрика (одно число)
                if widget_type == 'metric':
                    metric_data = self.calculate_metric(widget_config, filters)
                    widgets.append(metric_data)

                # Графики Chart.js (line, bar, pie, doughnut, radar, etc.)
                elif widget_type in ['line', 'bar', 'horizontalBar', 'pie', 'doughnut', 'polarArea', 'radar']:
                    chart_data = self.calculate_chart(widget_config, filters)
                    widgets.append(chart_data)

                # Таблица
                elif widget_type == 'table':
                    table_data = self.calculate_table(widget_config, filters)
                    widgets.append(table_data)

                # Старый формат: type='chart' с chart_type внутри
                elif widget_type == 'chart':
                    chart_data = self.calculate_chart(widget_config, filters)
                    widgets.append(chart_data)

            except Exception as e:
                # Пропустить виджеты с ошибками
                widgets.append({
                    'type': 'error',
                    'title': widget_config.get('title', 'Виджет'),
                    'error': str(e)
                })

        context['widgets'] = widgets
        context['filters'] = filters
        context['countries'] = Country.objects.all()
        context['regions'] = Region.objects.all()
        context['cities'] = City.objects.all()
        context['districts'] = District.objects.all()
        context['channels'] = Channel.objects.all()
        context['outlets'] = Outlet.objects.select_related('channel__district__city__region__country').all()
        return context

    def get_filters(self):
        """Получить и обработать фильтры"""
        period = self.request.GET.get('period', 'month')
        now = timezone.now()

        # Вычислить даты на основе периода
        if period == 'today':
            date_from = now.replace(hour=0, minute=0, second=0, microsecond=0)
            date_to = now
        elif period == 'yesterday':
            date_from = (now - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
            date_to = now.replace(hour=0, minute=0, second=0, microsecond=0)
        elif period == 'week':
            date_from = now - timedelta(days=now.weekday())
            date_to = now
        elif period == 'last_week':
            date_from = now - timedelta(days=now.weekday() + 7)
            date_to = now - timedelta(days=now.weekday())
        elif period == 'month':
            date_from = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            date_to = now
        elif period == 'custom':
            date_from_str = self.request.GET.get('date_from')
            date_to_str = self.request.GET.get('date_to')
            if date_from_str and date_to_str:
                from datetime import datetime
                date_from = timezone.make_aware(datetime.strptime(date_from_str, '%Y-%m-%d'))
                date_to = timezone.make_aware(datetime.strptime(date_to_str, '%Y-%m-%d'))
            else:
                date_from = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                date_to = now
        else:
            date_from = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            date_to = now

        return {
            'period': period,
            'date_from': date_from,
            'date_to': date_to,
            'country': self.request.GET.get('country'),
            'region': self.request.GET.get('region'),
            'city': self.request.GET.get('city'),
            'district': self.request.GET.get('district'),
            'channel': self.request.GET.get('channel'),
            'outlet': self.request.GET.get('outlet'),
            'data_type': self.request.GET.get('data_type', 'MON'),  # По умолчанию мониторинговые
        }

    def calculate_metric(self, config, filters):
        """Вычислить числовую метрику"""
        coefficient_id = config.get('coefficient_id')
        aggregation = config.get('aggregation', 'avg')

        # ВАЖНО: Тип данных берется из фильтров (выбран пользователем на дашборде)
        data_type = filters.get('data_type', 'MON')  # MON/EXP/AI

        # Базовый queryset с фильтрами
        queryset = Observation.objects.filter(
            visit__start_date__gte=filters['date_from'],
            visit__start_date__lte=filters['date_to']
        )

        # Применить дополнительные фильтры
        if filters.get('country'):
            queryset = queryset.filter(visit__outlet__channel__district__city__region__country_id=filters['country'])
        if filters.get('region'):
            queryset = queryset.filter(visit__outlet__channel__district__city__region_id=filters['region'])
        if filters.get('city'):
            queryset = queryset.filter(visit__outlet__channel__district__city_id=filters['city'])
        if filters.get('district'):
            queryset = queryset.filter(visit__outlet__channel__district_id=filters['district'])
        if filters.get('channel'):
            queryset = queryset.filter(visit__outlet__channel_id=filters['channel'])
        if filters.get('outlet'):
            queryset = queryset.filter(visit__outlet_id=filters['outlet'])

        # Фильтр по атрибутам
        if filters.get('attributes'):
            for attr_code, attr_value in filters['attributes'].items():
                if attr_value:
                    queryset = queryset.filter(
                        product__attribute_values__attribute__code=attr_code
                    )

        # Фильтр по коэффициенту
        if coefficient_id:
            queryset = queryset.filter(coefficient_id=coefficient_id)

        # ВАЖНО: Фильтр по типу источника данных
        if data_type in ['MON', 'EXP', 'AI']:
            queryset = queryset.filter(data_source_type=data_type)

        # Вычислить агрегацию
        value = 0
        if aggregation == 'avg':
            result = queryset.aggregate(Avg('value_numeric'))
            value = result['value_numeric__avg'] or 0
        elif aggregation == 'sum':
            result = queryset.aggregate(Sum('value_numeric'))
            value = result['value_numeric__sum'] or 0
        elif aggregation == 'count':
            value = queryset.count()
        elif aggregation == 'min':
            result = queryset.aggregate(Min('value_numeric'))
            value = result['value_numeric__min'] or 0
        elif aggregation == 'max':
            result = queryset.aggregate(Max('value_numeric'))
            value = result['value_numeric__max'] or 0

        return {
            'type': 'metric',
            'title': config.get('title', 'Метрика'),
            'value': round(float(value), 2) if value else 0,
            'unit': config.get('unit', ''),
            'color': config.get('color', 'primary'),
        }

    def calculate_chart(self, config, filters):
        """Подготовить данные для графика"""
        coefficient_id = config.get('coefficient_id')
        widget_type = config.get('type')
        chart_type = config.get('chart_type', widget_type or 'line')
        group_by = config.get('group_by', 'date')

        # ВАЖНО: Тип данных берется из фильтров (выбран пользователем на дашборде)
        data_type = filters.get('data_type', 'MON')  # MON/EXP/AI

        # Базовый queryset
        queryset = Observation.objects.filter(
            visit__start_date__gte=filters['date_from'],
            visit__start_date__lte=filters['date_to']
        )

        if filters.get('country'):
            queryset = queryset.filter(visit__outlet__channel__district__city__region__country_id=filters['country'])
        if filters.get('region'):
            queryset = queryset.filter(visit__outlet__channel__district__city__region_id=filters['region'])
        if filters.get('city'):
            queryset = queryset.filter(visit__outlet__channel__district__city_id=filters['city'])
        if filters.get('district'):
            queryset = queryset.filter(visit__outlet__channel__district_id=filters['district'])
        if filters.get('channel'):
            queryset = queryset.filter(visit__outlet__channel_id=filters['channel'])
        if filters.get('outlet'):
            queryset = queryset.filter(visit__outlet_id=filters['outlet'])

        # Фильтр по атрибутам
        if filters.get('attributes'):
            for attr_code, attr_value in filters['attributes'].items():
                if attr_value:
                    queryset = queryset.filter(
                        product__attribute_values__attribute__code=attr_code
                    )

        if coefficient_id:
            queryset = queryset.filter(coefficient_id=coefficient_id)

        # ВАЖНО: Фильтр по типу источника данных
        if data_type in ['MON', 'EXP', 'AI']:
            queryset = queryset.filter(data_source_type=data_type)

        labels = []
        values = []

        # Группировка по датам
        if group_by == 'date':
            data = queryset.extra(
                select={'date': "DATE(created_at)"}
            ).values('date').annotate(
                avg_value=Avg('value_numeric')
            ).order_by('date')[:30]  # Последние 30 точек

            labels = [item['date'].strftime('%d.%m') for item in data]
            values = [float(item['avg_value']) if item['avg_value'] else 0 for item in data]

        # Группировка по неделям
        elif group_by == 'week':
            data = queryset.extra(
                select={'week': "DATE_FORMAT(created_at, '%%Y-%%u')"}
            ).values('week').annotate(
                avg_value=Avg('value_numeric')
            ).order_by('week')[:12]

            labels = [f"Неделя {item['week']}" for item in data]
            values = [float(item['avg_value']) if item['avg_value'] else 0 for item in data]

        # Группировка по месяцам
        elif group_by == 'month':
            data = queryset.extra(
                select={'month': "DATE_FORMAT(created_at, '%%Y-%%m')"}
            ).values('month').annotate(
                avg_value=Avg('value_numeric')
            ).order_by('month')[:12]

            labels = [item['month'] for item in data]
            values = [float(item['avg_value']) if item['avg_value'] else 0 for item in data]

        # Группировка по регионам
        elif group_by == 'region':
            max_segments = config.get('max_segments', 10)
            data = queryset.select_related(
                'visit__outlet__channel__district__city__region'
            ).values(
                'visit__outlet__channel__district__city__region__name'
            ).annotate(
                avg_value=Avg('value_numeric')
            ).order_by('-avg_value')[:max_segments]

            labels = [item['visit__outlet__channel__district__city__region__name'] or 'Без региона' for item in data]
            values = [float(item['avg_value']) if item['avg_value'] else 0 for item in data]

        # Группировка по каналам
        elif group_by == 'channel':
            max_segments = config.get('max_segments', 10)
            data = queryset.select_related(
                'visit__outlet__channel'
            ).values(
                'visit__outlet__channel__name'
            ).annotate(
                avg_value=Avg('value_numeric')
            ).order_by('-avg_value')[:max_segments]

            labels = [item['visit__outlet__channel__name'] or 'Без канала' for item in data]
            values = [float(item['avg_value']) if item['avg_value'] else 0 for item in data]

        # Группировка по торговым точкам
        elif group_by == 'outlet':
            max_segments = config.get('max_segments', 10)
            data = queryset.select_related(
                'visit__outlet'
            ).values(
                'visit__outlet__name'
            ).annotate(
                avg_value=Avg('value_numeric')
            ).order_by('-avg_value')[:max_segments]

            labels = [item['visit__outlet__name'] or 'Без точки' for item in data]
            values = [float(item['avg_value']) if item['avg_value'] else 0 for item in data]

        return {
            'type': 'chart',
            'title': config.get('title', 'График'),
            'chart_type': chart_type,
            'labels': json.dumps(labels),
            'values': json.dumps(values),
            'color': config.get('color', 'rgba(75, 192, 192, 0.8)'),
        }

    def calculate_table(self, config, filters):
        """Подготовить данные для таблицы"""
        coefficient_id = config.get('coefficient_id')
        group_by = config.get('group_by', 'outlet')
        row_limit = config.get('row_limit', 10)
        sort_order = config.get('sort', 'desc')

        # ВАЖНО: Тип данных берется из фильтров
        data_type = filters.get('data_type', 'MON')

        # Базовый queryset
        queryset = Observation.objects.filter(
            visit__start_date__gte=filters['date_from'],
            visit__start_date__lte=filters['date_to']
        )

        if filters.get('country'):
            queryset = queryset.filter(visit__outlet__channel__district__city__region__country_id=filters['country'])
        if filters.get('region'):
            queryset = queryset.filter(visit__outlet__channel__district__city__region_id=filters['region'])
        if filters.get('city'):
            queryset = queryset.filter(visit__outlet__channel__district__city_id=filters['city'])
        if filters.get('district'):
            queryset = queryset.filter(visit__outlet__channel__district_id=filters['district'])
        if filters.get('channel'):
            queryset = queryset.filter(visit__outlet__channel_id=filters['channel'])
        if filters.get('outlet'):
            queryset = queryset.filter(visit__outlet_id=filters['outlet'])

        # Фильтр по атрибутам
        if filters.get('attributes'):
            for attr_code, attr_value in filters['attributes'].items():
                if attr_value:
                    queryset = queryset.filter(
                        product__attribute_values__attribute__code=attr_code
                    )

        if coefficient_id:
            queryset = queryset.filter(coefficient_id=coefficient_id)

        if data_type in ['MON', 'EXP', 'AI']:
            queryset = queryset.filter(data_source_type=data_type)

        # Получить данные
        rows = []
        order_by_clause = '-avg_value' if sort_order == 'desc' else 'avg_value'

        if group_by == 'outlet':
            data = queryset.select_related(
                'visit__outlet'
            ).values(
                'visit__outlet__name', 'visit__outlet__code'
            ).annotate(
                avg_value=Avg('value_numeric'),
                count=Count('id')
            ).order_by(order_by_clause)[:row_limit]

            rows = [
                {
                    'name': item['visit__outlet__name'] or 'Без названия',
                    'code': item['visit__outlet__code'] or '-',
                    'value': round(float(item['avg_value']), 2) if item['avg_value'] else 0,
                    'count': item['count']
                }
                for item in data
            ]

        elif group_by == 'region':
            data = queryset.select_related(
                'visit__outlet__channel__district__city__region'
            ).values(
                'visit__outlet__channel__district__city__region__name'
            ).annotate(
                avg_value=Avg('value_numeric'),
                count=Count('id')
            ).order_by(order_by_clause)[:row_limit]

            rows = [
                {
                    'name': item['visit__outlet__channel__district__city__region__name'] or 'Без региона',
                    'value': round(float(item['avg_value']), 2) if item['avg_value'] else 0,
                    'count': item['count']
                }
                for item in data
            ]

        elif group_by == 'channel':
            data = queryset.select_related(
                'visit__outlet__channel'
            ).values(
                'visit__outlet__channel__name'
            ).annotate(
                avg_value=Avg('value_numeric'),
                count=Count('id')
            ).order_by(order_by_clause)[:row_limit]

            rows = [
                {
                    'name': item['visit__outlet__channel__name'] or 'Без канала',
                    'value': round(float(item['avg_value']), 2) if item['avg_value'] else 0,
                    'count': item['count']
                }
                for item in data
            ]

        elif group_by == 'date':
            data = queryset.extra(
                select={'date': "DATE(created_at)"}
            ).values('date').annotate(
                avg_value=Avg('value_numeric'),
                count=Count('id')
            ).order_by(order_by_clause)[:row_limit]

            rows = [
                {
                    'name': item['date'].strftime('%d.%m.%Y'),
                    'value': round(float(item['avg_value']), 2) if item['avg_value'] else 0,
                    'count': item['count']
                }
                for item in data
            ]

        return {
            'type': 'table',
            'title': config.get('title', 'Таблица'),
            'rows': rows,
            'group_by': group_by,
        }


class MultiLevelDashboardView(LoginRequiredMixin, TemplateView):
    """
    Мультиуровневый дашборд с горизонтальным переключателем уровней:
    - Динамически загружает все дашборды с заполненным полем level из БД
    - Автоматически добавляет/удаляет уровни при изменении дашбордов
    """
    template_name = 'analytics/multilevel_dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Получить все активные дашборды с уровнями из БД
        level_dashboards = Dashboard.objects.filter(
            is_active=True,
            level__isnull=False
        ).exclude(level='').order_by('level_order')

        # Если нет дашбордов с уровнями, показать ошибку
        if not level_dashboards.exists():
            context['error'] = 'Нет доступных дашбордов с уровнями'
            context['widgets'] = []
            return context

        # Получить текущий уровень из GET-параметра
        level = self.request.GET.get('level')

        # Если уровень не указан, взять первый доступный
        if not level:
            dashboard = level_dashboards.first()
            level = dashboard.level
        else:
            # Найти дашборд для выбранного уровня
            try:
                dashboard = level_dashboards.get(level=level)
            except Dashboard.DoesNotExist:
                # Если дашборд для этого уровня не найден, взять первый
                dashboard = level_dashboards.first()
                level = dashboard.level

        context['level'] = level
        context['dashboard'] = dashboard
        context['level_dashboards'] = level_dashboards  # Все доступные дашборды для переключателя

        # Получить фильтры
        filters = self.get_filters()
        context['filters'] = filters

        # Данные для селекторов
        context['countries'] = Country.objects.all()
        context['regions'] = Region.objects.all()
        context['cities'] = City.objects.all()
        context['districts'] = District.objects.all()
        context['channels'] = Channel.objects.all()
        context['outlets'] = Outlet.objects.select_related('channel__district__city__region__country').all()

        # Загрузить группы атрибутов для фильтрации
        context['attribute_groups'] = AttributeGroup.objects.prefetch_related(
            'attributes'
        ).filter(
            attributes__is_filterable=True
        ).distinct().order_by('order')

        # Вычислить виджеты дашборда
        widgets_config = dashboard.widgets_config if dashboard.widgets_config else {}
        widgets = []

        for widget_config in widgets_config.get('widgets', []):
            try:
                widget_type = widget_config.get('type')

                if widget_type == 'metric':
                    widget_data = self.calculate_metric(widget_config, filters)
                    widgets.append(widget_data)
                elif widget_type == 'chart':
                    widget_data = self.calculate_chart(widget_config, filters)
                    widgets.append(widget_data)
                elif widget_type == 'table':
                    widget_data = self.calculate_table(widget_config, filters)
                    widgets.append(widget_data)
            except Exception as e:
                widgets.append({
                    'type': 'error',
                    'title': widget_config.get('title', 'Виджет'),
                    'error': str(e)
                })

        context['widgets'] = widgets
        return context

    def get_filters(self):
        """Получить и обработать фильтры"""
        period = self.request.GET.get('period', 'month')
        now = timezone.now()

        # Вычислить даты на основе периода
        if period == 'today':
            date_from = now.replace(hour=0, minute=0, second=0, microsecond=0)
            date_to = now
        elif period == 'yesterday':
            date_from = (now - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
            date_to = date_from.replace(hour=23, minute=59, second=59)
        elif period == 'week':
            date_from = (now - timedelta(days=now.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
            date_to = now
        elif period == 'last_week':
            date_from = (now - timedelta(days=now.weekday() + 7)).replace(hour=0, minute=0, second=0, microsecond=0)
            date_to = date_from + timedelta(days=6, hours=23, minutes=59, seconds=59)
        elif period == 'custom':
            date_from_str = self.request.GET.get('date_from')
            date_to_str = self.request.GET.get('date_to')
            if date_from_str and date_to_str:
                date_from = timezone.make_aware(timezone.datetime.strptime(date_from_str, '%Y-%m-%d'))
                date_to = timezone.make_aware(timezone.datetime.strptime(date_to_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59))
            else:
                date_from = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                date_to = now
        else:
            date_from = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            date_to = now

        # Получить фильтры по уровню
        level = self.request.GET.get('level', 'country')
        entity_id = self.request.GET.get('entity_id')

        # Получить фильтры по атрибутам (все параметры, начинающиеся с 'attr_')
        attribute_filters = {}
        for key, value in self.request.GET.items():
            if key.startswith('attr_') and value:
                attr_code = key[5:]  # Убрать префикс 'attr_'
                attribute_filters[attr_code] = value

        return {
            'period': period,
            'date_from': date_from,
            'date_to': date_to,
            'level': level,
            'country': entity_id if level == 'country' else None,
            'region': entity_id if level == 'region' else None,
            'city': entity_id if level == 'city' else None,
            'district': entity_id if level == 'district' else None,
            'channel': entity_id if level == 'channel' else None,
            'outlet': entity_id if level == 'outlet' else None,
            'data_type': self.request.GET.get('data_type', 'MON'),
            'attributes': attribute_filters,
        }

    # Используем те же методы calculate_metric, calculate_chart, calculate_table
    # что и в DashboardDetailView
    calculate_metric = DashboardDetailView.calculate_metric
    calculate_chart = DashboardDetailView.calculate_chart
    calculate_table = DashboardDetailView.calculate_table


class DashboardCreateView(LoginRequiredMixin, CreateView):
    model = Dashboard
    template_name = 'analytics/dashboard_form.html'
    form_class = DashboardForm
    success_url = reverse_lazy('analytics:dashboard_list')

    def form_valid(self, form):
        form.instance.owner = self.request.user
        messages.success(self.request, f'Dashboard "{form.instance.name}" created successfully.')
        return super().form_valid(form)


class DashboardUpdateView(LoginRequiredMixin, UpdateView):
    model = Dashboard
    template_name = 'analytics/dashboard_form.html'
    form_class = DashboardForm
    success_url = reverse_lazy('analytics:dashboard_list')

    def form_valid(self, form):
        messages.success(self.request, f'Dashboard "{form.instance.name}" updated successfully.')
        return super().form_valid(form)


class DashboardDeleteView(LoginRequiredMixin, DeleteView):
    model = Dashboard
    template_name = 'analytics/dashboard_confirm_delete.html'
    success_url = reverse_lazy('analytics:dashboard_list')

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, f'Dashboard "{self.get_object().name}" deleted successfully.')
        return super().delete(request, *args, **kwargs)


# ============================================================================
# Report Views
# ============================================================================

class ReportListView(LoginRequiredMixin, ListView):
    model = Report
    template_name = 'analytics/report_list.html'
    context_object_name = 'reports'
    paginate_by = 25

    def get_queryset(self):
        return Report.objects.select_related('template', 'created_by')


class ReportDetailView(LoginRequiredMixin, DetailView):
    model = Report
    template_name = 'analytics/report_detail.html'
    context_object_name = 'report'

    def get_queryset(self):
        return Report.objects.select_related('template', 'created_by')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Парсить result JSON и подготовить для отображения
        result = {}
        if self.object.result:
            try:
                result = json.loads(self.object.result) if isinstance(self.object.result, str) else self.object.result
            except:
                result = {}

        # Преобразовать в структурированные данные
        tables = []
        charts = []
        metrics = []

        for section_key, section_data in result.items():
            if isinstance(section_data, list) and len(section_data) > 0:
                # Это таблица
                tables.append({
                    'title': section_key.replace('_', ' ').title(),
                    'data': section_data
                })
            elif isinstance(section_data, dict):
                if 'labels' in section_data and 'values' in section_data:
                    # Это график
                    charts.append({
                        'title': section_key.replace('_', ' ').title(),
                        'type': section_data.get('type', 'bar'),
                        'labels': json.dumps(section_data.get('labels', [])),
                        'values': json.dumps(section_data.get('values', []))
                    })
                else:
                    # Это метрика
                    metrics.append({
                        'title': section_key.replace('_', ' ').title(),
                        'value': section_data.get('value', section_data),
                        'unit': section_data.get('unit', '')
                    })
            elif isinstance(section_data, (int, float)):
                # Простое число
                metrics.append({
                    'title': section_key.replace('_', ' ').title(),
                    'value': section_data,
                    'unit': ''
                })

        context['tables'] = tables
        context['charts'] = charts
        context['metrics'] = metrics
        return context


class ReportCreateView(LoginRequiredMixin, CreateView):
    model = Report
    template_name = 'analytics/report_form.html'
    form_class = ReportForm
    success_url = reverse_lazy('analytics:report_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, f'Report "{form.instance.name}" created successfully.')
        return super().form_valid(form)


class ReportUpdateView(LoginRequiredMixin, UpdateView):
    model = Report
    template_name = 'analytics/report_form.html'
    form_class = ReportForm
    success_url = reverse_lazy('analytics:report_list')

    def form_valid(self, form):
        messages.success(self.request, f'Report "{form.instance.name}" updated successfully.')
        return super().form_valid(form)


class ReportDeleteView(LoginRequiredMixin, DeleteView):
    model = Report
    template_name = 'analytics/report_confirm_delete.html'
    success_url = reverse_lazy('analytics:report_list')

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, f'Report "{self.get_object().name}" deleted successfully.')
        return super().delete(request, *args, **kwargs)


class ReportPDFView(LoginRequiredMixin, DetailView):
    """Generate PDF version of report"""
    model = Report

    def get(self, request, *args, **kwargs):
        # Попытка импортировать WeasyPrint только когда нужно
        try:
            from weasyprint import HTML
        except (ImportError, OSError) as e:
            messages.error(request, 'PDF generation is not available. WeasyPrint library or GTK dependencies are not installed.')
            return HttpResponse('PDF generation not available. Please install WeasyPrint and GTK dependencies.', status=500)

        report = self.get_object()

        # Подготовить данные для отчета
        result = {}
        if report.result:
            try:
                result = json.loads(report.result) if isinstance(report.result, str) else report.result
            except:
                result = {}

        # Преобразовать в структурированные данные (как в ReportDetailView)
        tables = []
        metrics = []

        for section_key, section_data in result.items():
            if isinstance(section_data, list) and len(section_data) > 0:
                tables.append({
                    'title': section_key.replace('_', ' ').title(),
                    'data': section_data
                })
            elif isinstance(section_data, dict):
                if 'value' in section_data or not ('labels' in section_data and 'values' in section_data):
                    metrics.append({
                        'title': section_key.replace('_', ' ').title(),
                        'value': section_data.get('value', section_data),
                        'unit': section_data.get('unit', '')
                    })
            elif isinstance(section_data, (int, float)):
                metrics.append({
                    'title': section_key.replace('_', ' ').title(),
                    'value': section_data,
                    'unit': ''
                })

        context = {
            'report': report,
            'tables': tables,
            'metrics': metrics,
            'generated_at': timezone.now(),
        }

        # Рендерить HTML шаблон
        html_string = render_to_string('analytics/report_pdf.html', context)

        # Конвертировать в PDF
        html = HTML(string=html_string, base_url=request.build_absolute_uri())
        pdf = html.write_pdf()

        # Вернуть PDF
        response = HttpResponse(pdf, content_type='application/pdf')
        filename = f'report_{report.code}_{timezone.now().strftime("%Y%m%d_%H%M")}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class ReportExcelView(LoginRequiredMixin, DetailView):
    """Export report to Excel format"""
    model = Report

    def get(self, request, *args, **kwargs):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messages.error(request, 'Excel export is not available. Please install openpyxl.')
            return HttpResponse('Excel export not available. Please install openpyxl.', status=500)

        report = self.get_object()

        # Подготовить данные для отчета
        result = {}
        if report.result:
            try:
                result = json.loads(report.result) if isinstance(report.result, str) else report.result
            except:
                result = {}

        # Создать workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"

        # Заголовок отчета
        ws['A1'] = report.name
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = report.description or ''
        ws['A2'].font = Font(size=11, italic=True)

        ws['A3'] = f"Дата создания: {report.created_at.strftime('%d.%m.%Y %H:%M')}"
        ws['A4'] = f"Статус: {report.get_status_display()}"

        row = 6  # Начать с 6й строки

        # Добавить метрики
        metrics = []
        for section_key, section_data in result.items():
            if isinstance(section_data, dict) and 'value' in section_data:
                metrics.append({
                    'title': section_key.replace('_', ' ').title(),
                    'value': section_data.get('value', section_data),
                    'unit': section_data.get('unit', '')
                })
            elif isinstance(section_data, (int, float)):
                metrics.append({
                    'title': section_key.replace('_', ' ').title(),
                    'value': section_data,
                    'unit': ''
                })

        if metrics:
            ws[f'A{row}'] = 'МЕТРИКИ'
            ws[f'A{row}'].font = Font(bold=True, size=14)
            ws[f'A{row}'].fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            row += 1

            for metric in metrics:
                ws[f'A{row}'] = metric['title']
                ws[f'B{row}'] = f"{metric['value']} {metric['unit']}"
                row += 1

            row += 2

        # Добавить таблицы
        for section_key, section_data in result.items():
            if isinstance(section_data, list) and len(section_data) > 0:
                # Заголовок секции
                ws[f'A{row}'] = section_key.replace('_', ' ').title()
                ws[f'A{row}'].font = Font(bold=True, size=14)
                ws[f'A{row}'].fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                row += 1

                # Заголовки столбцов
                headers = list(section_data[0].keys())
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=row, column=col_idx, value=header.title())
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center')

                row += 1

                # Данные
                for data_row in section_data:
                    for col_idx, value in enumerate(data_row.values(), start=1):
                        ws.cell(row=row, column=col_idx, value=str(value))
                    row += 1

                row += 2

        # Автоширина столбцов
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Сохранить в response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f'report_{report.code}_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


class ReportCSVView(LoginRequiredMixin, DetailView):
    """Export report to CSV format"""
    model = Report

    def get(self, request, *args, **kwargs):
        import csv

        report = self.get_object()

        # Подготовить данные для отчета
        result = {}
        if report.result:
            try:
                result = json.loads(report.result) if isinstance(report.result, str) else report.result
            except:
                result = {}

        # Создать CSV response
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        filename = f'report_{report.code}_{timezone.now().strftime("%Y%m%d_%H%M")}.csv'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # BOM для правильного отображения в Excel
        response.write('\ufeff')

        writer = csv.writer(response)

        # Заголовок отчета
        writer.writerow([report.name])
        writer.writerow([report.description or ''])
        writer.writerow([f"Дата создания: {report.created_at.strftime('%d.%m.%Y %H:%M')}"])
        writer.writerow([f"Статус: {report.get_status_display()}"])
        writer.writerow([])

        # Метрики
        metrics = []
        for section_key, section_data in result.items():
            if isinstance(section_data, dict) and 'value' in section_data:
                metrics.append({
                    'title': section_key.replace('_', ' ').title(),
                    'value': section_data.get('value', section_data),
                    'unit': section_data.get('unit', '')
                })
            elif isinstance(section_data, (int, float)):
                metrics.append({
                    'title': section_key.replace('_', ' ').title(),
                    'value': section_data,
                    'unit': ''
                })

        if metrics:
            writer.writerow(['МЕТРИКИ'])
            for metric in metrics:
                writer.writerow([metric['title'], f"{metric['value']} {metric['unit']}"])
            writer.writerow([])

        # Таблицы
        for section_key, section_data in result.items():
            if isinstance(section_data, list) and len(section_data) > 0:
                # Заголовок секции
                writer.writerow([section_key.replace('_', ' ').title()])

                # Заголовки столбцов
                headers = list(section_data[0].keys())
                writer.writerow([h.title() for h in headers])

                # Данные
                for data_row in section_data:
                    writer.writerow(list(data_row.values()))

                writer.writerow([])

        return response


# ============================================================================
# ReportTemplate Views
# ============================================================================

class ReportTemplateListView(LoginRequiredMixin, ListView):
    model = ReportTemplate
    template_name = 'analytics/reporttemplate_list.html'
    context_object_name = 'reporttemplates'
    paginate_by = 25

    def get_queryset(self):
        return ReportTemplate.objects.select_related('created_by')


class ReportTemplateDetailView(LoginRequiredMixin, DetailView):
    model = ReportTemplate
    template_name = 'analytics/reporttemplate_detail.html'
    context_object_name = 'reporttemplate'

    def get_queryset(self):
        return ReportTemplate.objects.select_related('created_by').prefetch_related('metrics', 'reports')


class ReportTemplateCreateView(LoginRequiredMixin, CreateView):
    model = ReportTemplate
    template_name = 'analytics/reporttemplate_form.html'
    form_class = ReportTemplateForm
    success_url = reverse_lazy('analytics:reporttemplate_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, f'Report Template "{form.instance.name}" created successfully.')
        return super().form_valid(form)


class ReportTemplateUpdateView(LoginRequiredMixin, UpdateView):
    model = ReportTemplate
    template_name = 'analytics/reporttemplate_form.html'
    form_class = ReportTemplateForm
    success_url = reverse_lazy('analytics:reporttemplate_list')

    def form_valid(self, form):
        messages.success(self.request, f'Report Template "{form.instance.name}" updated successfully.')
        return super().form_valid(form)


class ReportTemplateDeleteView(LoginRequiredMixin, DeleteView):
    model = ReportTemplate
    template_name = 'analytics/reporttemplate_confirm_delete.html'
    success_url = reverse_lazy('analytics:reporttemplate_list')

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, f'Report Template "{self.get_object().name}" deleted successfully.')
        return super().delete(request, *args, **kwargs)


# ============================================================================
# FilterPreset Views
# ============================================================================

class FilterPresetListView(LoginRequiredMixin, ListView):
    model = FilterPreset
    template_name = 'analytics/filterpreset_list.html'
    context_object_name = 'filterpresets'
    paginate_by = 25

    def get_queryset(self):
        return FilterPreset.objects.select_related('owner')


class FilterPresetDetailView(LoginRequiredMixin, DetailView):
    model = FilterPreset
    template_name = 'analytics/filterpreset_detail.html'
    context_object_name = 'filterpreset'

    def get_queryset(self):
        return FilterPreset.objects.select_related('owner')


class FilterPresetCreateView(LoginRequiredMixin, CreateView):
    model = FilterPreset
    template_name = 'analytics/filterpreset_form.html'
    form_class = FilterPresetForm
    success_url = reverse_lazy('analytics:filterpreset_list')

    def form_valid(self, form):
        form.instance.owner = self.request.user
        messages.success(self.request, f'Filter Preset "{form.instance.name}" created successfully.')
        return super().form_valid(form)


class FilterPresetUpdateView(LoginRequiredMixin, UpdateView):
    model = FilterPreset
    template_name = 'analytics/filterpreset_form.html'
    form_class = FilterPresetForm
    success_url = reverse_lazy('analytics:filterpreset_list')

    def form_valid(self, form):
        messages.success(self.request, f'Filter Preset "{form.instance.name}" updated successfully.')
        return super().form_valid(form)


class FilterPresetDeleteView(LoginRequiredMixin, DeleteView):
    model = FilterPreset
    template_name = 'analytics/filterpreset_confirm_delete.html'
    success_url = reverse_lazy('analytics:filterpreset_list')

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, f'Filter Preset "{self.get_object().name}" deleted successfully.')
        return super().delete(request, *args, **kwargs)
